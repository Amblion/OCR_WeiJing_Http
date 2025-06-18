#coding=utf-8
import sys
import threading
import os
import io
import base64
from PIL import Image as imim
import requests
from os.path import abspath, dirname
from openpyxl import  Workbook 
from openpyxl.styles import Alignment
from tkinter import *
import Fun
import chardet
import smbclient
sys.path.insert(0,abspath(dirname(__file__)))
uiName="OCR_WeiJing_Http"
ElementBGArray={}  
ElementBGArray_Resize={} 
ElementBGArray_IM={} 
class Dnconsole:
    '''
    【雷电控制台类】
    version: 9.0
    import该文件会自动实例化为 Dc
    '''
    def __init__( self, installation_path:str ):
        '''
        【构造方法】
        '''
        # if 模拟器安装路径存在性检测
        if os.path.exists(installation_path) is False:
            print('模拟器安装路径不存在！')
        # 获取模拟器安装路径
        self.ins_path = installation_path
        # Dnconsole程序路径
        self.console_path = self.ins_path + r'\ldconsole.exe '
        # if Dnconsole程序路径检测
        if os.path.exists(self.console_path) is False:
            print('Dnconsole程序路径不存在！\n请确认模拟器安装文件是否完整或模拟器版本是否不符！')
        # adb程序路径
        self.adb_path = self.ins_path + r'\adb.exe '
        # if adb程序路径检测
        if os.path.exists(self.adb_path) is False:
            print('Dnconsole程序路径不存在！\n请确认模拟器安装文件是否完整！')
        # 模拟器截屏程序路径
        self.screencap_path = r'/system/bin/screencap'
        # 模拟器截图保存路径
        self.devicess_path = r'/sdcard/autosS.png'
        # 本地图片保存路径
        self.images_path = r'C:\leidian\images'
        # 构造完成
        print('Class-Dnconsole is ready.(%s)' % (self.ins_path))
    def CMD( self, cmd:str ):
        '''
        【执行控制台命令语句】
        :param cmd: 命令
        :return: 控制台调试内容
        '''
        CMD = self.console_path + cmd # 控制台命令
        process = os.popen(CMD)
        result = process.read()
        process.close()
        return result
    def ADB( self, cmd:str ):
        '''
        【执行ADB命令语句】
        :param cmd: 命令
        :return: 控制台调试内容
        '''
        CMD = self.adb_path + cmd # adb命令
        process = os.popen(CMD)
        result = process.read()
        process.close()
        return result    
    def list(self):
        '''
        【获取模拟器列表（仅标题）】
        :return: 控制台调试内容
        '''
        cmd = 'list'
        return self.CMD(cmd)     
    def list2(self):
        '''
        【取模拟器列表】
        :return: 列表（索引、标题、顶层句柄、绑定句柄、是否进入android、进程PID、VBox进程PID）
        '''
        cmd = 'list2'
        return self.CMD(cmd)
    def runninglist(self):
        '''
        【获取正在运行的模拟器列表（仅标题）】
        :return: 控制台调试内容
        '''
        cmd = 'runninglist'
        return self.CMD(cmd)
    def screenShot ( self, index:int ):
        '''
        【截屏并保存到本地】
        :param index: 模拟器序号
        '''
        path = "C:\\leidian\\%s.png"%(index)
        cmd1 = 'adb --index %d --command "shell %s -p %s"' %(index, self.screencap_path, self.devicess_path)
        cmd2 = 'adb --index %d --command "pull %s %s"' %(index, self.devicess_path, r""+path)
        self.CMD(cmd1)
        self.CMD(cmd2)

def get_xlsx(data):
    # 实例化
    wb = Workbook()
    # 激活 worksheet
    ws = wb.create_sheet(title="数字数据",index=0)
    ws3 = wb.create_sheet(title="转换后数据",index=1)
    align = Alignment(horizontal='center', vertical='center')
     # 设置列宽
    ws.column_dimensions['A'].width = 20  # 设置第一列的宽度为 20
    ws.column_dimensions['B'].width = 15  # 设置第二列的宽度为 15
    ws.column_dimensions['C'].width = 15  # 设置第三列的宽度为 15
    ws.column_dimensions['D'].width = 15  # 设置第四列的宽度为 15
    ws.column_dimensions['E'].width = 10  # 设置第四列的宽度为 10
    ws.column_dimensions['F'].width = 15  # 设置第六列的宽度为 15
    ws.column_dimensions['G'].width = 15  # 设置第七列的宽度为 15
    ws.column_dimensions['H'].width = 15  # 设置第八列的宽度为 15
    ws.column_dimensions['I'].width = 10  # 设置第五列的宽度为 10
    ws.column_dimensions['J'].width = 15  # 设置第九列的宽度为 15
    ws.column_dimensions['K'].width = 15  # 设置第九列的宽度为 15
    ws.column_dimensions['L'].width = 15  # 设置第九列的宽度为 15
    ws.column_dimensions['M'].width = 10  # 设置第九列的宽度为 10
    ws.column_dimensions['N'].width = 15  # 设置第九列的宽度为 15
    ws.column_dimensions['O'].width = 15  # 设置第九列的宽度为 15
    ws.column_dimensions['P'].width = 15  # 设置第九列的宽度为 15
    ws3.column_dimensions['A'].width = 20  # 设置第一列的宽度为 20
    ws3.column_dimensions['B'].width = 15  # 设置第二列的宽度为 15
    ws3.column_dimensions['C'].width = 15  # 设置第三列的宽度为 15
    ws3.column_dimensions['D'].width = 15  # 设置第四列的宽度为 15
    ws3.column_dimensions['E'].width = 10  # 设置第四列的宽度为 10
    ws3.column_dimensions['F'].width = 15  # 设置第六列的宽度为 15
    ws3.column_dimensions['G'].width = 15  # 设置第七列的宽度为 15
    ws3.column_dimensions['H'].width = 15  # 设置第八列的宽度为 15
    ws3.column_dimensions['I'].width = 10  # 设置第五列的宽度为 10
    ws3.column_dimensions['J'].width = 15  # 设置第九列的宽度为 15
    ws3.column_dimensions['K'].width = 15  # 设置第九列的宽度为 15
    ws3.column_dimensions['L'].width = 15  # 设置第九列的宽度为 15
    ws3.column_dimensions['M'].width = 10  # 设置第九列的宽度为 10
    ws3.column_dimensions['N'].width = 15  # 设置第九列的宽度为 15
    ws3.column_dimensions['O'].width = 15  # 设置第九列的宽度为 15
    ws3.column_dimensions['P'].width = 15  # 设置第九列的宽度为 15
    ws.append(["模拟器","食物(角色)","食物(背包)","食物(总额)","","木材(角色)","木材(背包)","木材(总额)","","石材(角色)","石材(背包)","石材(总额)","","金币(角色)","金币(背包)","金币(总额)"])
    ws3.append(["模拟器","食物(角色)","食物(背包)","食物(总额)","","木材(角色)","木材(背包)","木材(总额)","","石材(角色)","石材(背包)","石材(总额)","","金币(角色)","金币(背包)","金币(总额)"])
    for i in range(1, ws.max_row + 1):
        for j in range(1, ws.max_column + 1):
            ws.cell(i, j).alignment = align  
            ws3.cell(i, j).alignment = align 
    xlsx_index = 2
    for data_list in data:
        ws["A"+str(xlsx_index)] = data_list[0] 
        ws3["A"+str(xlsx_index)] = data_list[0]
        ws["B"+str(xlsx_index)] = data_list[1]
        ws["C"+str(xlsx_index)] = data_list[2]  
        ws["D"+str(xlsx_index)] = data_list[3]
        ws3["B"+str(xlsx_index)] = int((data_list[1]/ 1000000) * 100) / 100
        ws3["C"+str(xlsx_index)] = int((data_list[2]/ 1000000) * 100) / 100
        ws3["D"+str(xlsx_index)] = int((data_list[3]/ 1000000) * 100) / 100
        ws["F"+str(xlsx_index)] = data_list[4]  
        ws["G"+str(xlsx_index)] = data_list[5]  
        ws["H"+str(xlsx_index)] = data_list[6]
        ws3["F"+str(xlsx_index)] = int((data_list[4]/ 1000000) * 100) / 100
        ws3["G"+str(xlsx_index)] = int((data_list[5]/ 1000000) * 100) / 100 
        ws3["H"+str(xlsx_index)] = int((data_list[6]/ 1000000) * 100) / 100
        ws["J"+str(xlsx_index)] = data_list[7]  
        ws["K"+str(xlsx_index)] = data_list[8]  
        ws["L"+str(xlsx_index)] = data_list[9]
        ws3["J"+str(xlsx_index)] = int((data_list[7]/ 1000000) * 100) / 100 
        ws3["K"+str(xlsx_index)] = int((data_list[8]/ 1000000) * 100) / 100  
        ws3["L"+str(xlsx_index)] = int((data_list[9]/ 1000000) * 100) / 100 
        ws["N"+str(xlsx_index)] = data_list[10]  
        ws["O"+str(xlsx_index)] = data_list[11]  
        ws["P"+str(xlsx_index)] = data_list[12]   
        ws3["N"+str(xlsx_index)] = int((data_list[10]/ 1000000) * 100) / 100   
        ws3["O"+str(xlsx_index)] = int((data_list[11]/ 1000000) * 100) / 100    
        ws3["P"+str(xlsx_index)] = int((data_list[12]/ 1000000) * 100) / 100      
        xlsx_index+=1     
    for i in range(1, ws.max_row + 1):
        for j in range(1, ws.max_column + 1):
            ws.cell(i, j).alignment = align  
            ws3.cell(i, j).alignment = align
    xlsx_name = Fun.SaveFile("导出数据",[('XLS File','*.xls')], os.path.abspath('.'),'','.xls')       
    try:
        wb.save(xlsx_name) 
        Fun.MessageBox("保存成功","信息","info",None)
    except Exception as Ex:
        Fun.MessageBox("保存失败","错误","info",None)
        print(Ex)
        
def onLoad_thread():
    global onLoad 
    onLoad = False
    Fun.SetRowBGColor(uiName,'ListView_1','even','lightblue')
    global Dc
    Dc = Dnconsole(r'D:\leidian\LDPlayer9')
    onLoad = True
    print("软件初始化完成！")
    pass
    
#Form 'Form_1's Load Event :
def Form_1_onLoad(uiName,threadings=0):
    path = os.getcwd()
    if str(path).find("192.168.6.218") != -1:
        Fun.MessageBox("不支持在192.168.6.218共享目录内直接打开","运行错误","error",None)
        sys.exit() #退出主程序
    username = '技术部'  # 用户名
    password = 'Aa123654'  # 密码
    smbclient.reset_connection_cache()
    smbclient.ClientConfig(username=username, password=password)
    try:
        with smbclient.open_file("\\192.168.6.218\\技术部\\工具验证设置(非专业勿动勿删)\\资源导出\\version.txt",mode='rb') as f:
            raw_data = f.read()        
            result = chardet.detect(raw_data)
            encoding = result["encoding"]
        with smbclient.open_file(r"\\192.168.6.218\\技术部\\工具验证设置(非专业勿动勿删)\\资源导出\\version.txt",encoding=encoding) as f:
            new_version = f.read().split("----")
    except:
        Fun.MessageBox("验证异常！联系技术处理","运行错误","error",None)
        sys.exit() #退出主程序
        
    if(int(new_version[0]) > int(version)):
        if int(new_version[1]) == 1:
            Fun.MessageBox("有新版本，需要更新","更新提醒","info",None)
            writeUpgrade("资源导出.exe")
        Fun.SetText(uiName,'Label_2',"有新版本")
        Fun.SetTextColor(uiName,'Label_2','#00B800')
    else:
        Fun.SetText(uiName,'Label_2',"最新版本")
    onLoad_ = threading.Thread(target=onLoad_thread)
    onLoad_.start()    
img_path = 'C:\\leidian\\'    
def get_rec(leidian_index,list_index):
    img_name = str(leidian_index)+".png"
    Dc.screenShot(int(leidian_index))
    path = img_path+img_name
    # 打开一个图片文件
    img = imim.open(path)
    # 将图片转换为RGB格式（如果需要）
    img = img.convert('RGB')
    # 将图片转换为字节数据
    img_byte_arr = io.BytesIO()
    img.save(img_byte_arr, format='PNG')
    encoded_img = base64.b64encode(img_byte_arr.getvalue()).decode('utf-8')
    # 打印或使用Base64编码的图片数据
    r =  requests.post('http://192.168.23.50:5000/OCR', data={'image_data': encoded_img,"img_name": img_name})
    模拟器 = [""]+r.json()
    for 模拟器_index in range(1,len(模拟器)):
        Fun.SetCellText(uiName,'ListView_1',list_index,模拟器_index,模拟器[模拟器_index])
    print("雷电模拟器-%s 完成"%(leidian_index))    

def one_thread(ld_list):
    for leidian_index in ld_list:
        try: 
            模拟器 = [""]+["雷电模拟器-%s"%(leidian_index)]+["0","0","0"]+["0","0","0"]+["0","0","0"]+["0","0","0"]
            columnTextList = Fun.GetColumnTextList(uiName,'ListView_1',1)
            try:
                sss = columnTextList.index("雷电模拟器-%s"%(leidian_index))
            except ValueError:
                sss = Fun.AddRowText(uiName,'ListView_1','end',模拟器)
            Fun.SetCellCheckBox(uiName,'ListView_1',sss,0,False)
            get_rec(leidian_index,sss)
        except Exception as Ex:
            print(Ex)
            print("雷电模拟器-%s错误，检查模拟器分辨率ROOT等设置"%(leidian_index))
fag = True         
def get_rec_onLoad(thread_img):
    for run_thread in thread_img:
        run_thread.join()  
    global fag
    fag = True
#Button 'Button_1' 's Command Event :
def Button_1_onCommand(uiName,widgetName,threadings=0):
    global fag
    if fag and onLoad:
        fag = False
        if not os.path.exists(img_path):
            os.makedirs(img_path)
        num = os.listdir(img_path)
        if len(num) != 0:
            for filename in os.listdir(img_path):
                file_path = os.path.join(img_path, filename)
                if os.path.isfile(file_path) or os.path.islink(file_path):
                    os.unlink(file_path)
        ld_list = []
        data = Dc.list2().split('\n')
        for i in data:
            if len(i.split(",")) != 1:
                if i.split(",")[5] != "-1":
                    ld_list.append(int(i.split(",")[0]))
        ld_list.sort()
        thread_img = []    
        thread_s = False 
        if thread_s:
            print("还没有写")
        else:
            run_thread = threading.Thread(target=one_thread, args=[ld_list])
            thread_img.append(run_thread)
            run_thread.start()
        thread_img_state = threading.Thread(target=get_rec_onLoad, args=[thread_img])
        thread_img_state.start()
    else: 
        Fun.MessageBox("程序初始化或者识别操作中ing","提示","info",None)   
#Button 'Button_2' 's Command Event :
def Button_2_onCommand(uiName,widgetName,threadings=0):
    listviewTextList = Fun.GetAllRowTextList(uiName,'ListView_1')
    data = []
    for list_data in listviewTextList:
        data.append(list_data[1:])
    get_xlsx(data)  
    pass
version = "2025042902"    
#编写bat脚本，删除旧程序，运行新程序
def writeUpgrade(exe_name):
    img_path = 'C:\\updata\\'
    if not os.path.exists(img_path):
        os.makedirs(img_path)
    b = open("C:\\updata\\upgrade.bat",'w')
    TempList = "@echo off\n"
    TempList += "if not exist "+r"\\192.168.6.218\\维京崛起\\其它\\工具版本热更新(非专业勿动勿删)\\资源导出\\" + exe_name + " exit \n"  #判断是否有新版本的程序，没有就退出更新。
    TempList += "echo 正在更新至最新版本...\n"
    TempList += "timeout /t 10 /nobreak\n"  #等待10秒
    TempList += "del " + os.path.realpath(exe_name) + "\n" #删除旧程序
    TempList += "copy "+r"\\192.168.6.218\\维京崛起\\其它\\工具版本热更新(非专业勿动勿删)\\资源导出\\" + exe_name + " " + exe_name + '\n' #复制新版本程序
    TempList += "echo 更新完成\n"
    TempList += "timeout /t 3 /nobreak\n"
    # TempList += "start " + exe_name   #"start 1.bat\n"
    TempList += "exit"
    b.write(TempList)
    b.close()
    #subprocess.Popen("C:\\updata\\upgrade.bat") #不显示cmd窗口
    os.system('start C:\\updata\\upgrade.bat')  #显示cmd窗口
    sys.exit() #退出主程序
#Button 'Button_3' 's Command Event :
def Button_3_onCommand(uiName,widgetName,threadings=0):
    with smbclient.open_file("\\192.168.6.218\\技术部\\工具验证设置(非专业勿动勿删)\\资源导出\\version.txt",mode='rb') as f:
        raw_data = f.read()        
        result = chardet.detect(raw_data)
        encoding = result["encoding"]
    with smbclient.open_file(r"\\192.168.6.218\\技术部\\工具验证设置(非专业勿动勿删)\\资源导出\\version.txt",encoding=encoding) as f:
        new_version = f.read().split("----")
    if (int(new_version[0]) > int(version)):
        writeUpgrade("资源导出.exe")
    pass
#ListView 'ListView_1's CellClicked Event :
def ListView_1_onCellClicked(uiName,widgetName,rowIndex,columnIndex,threadings=0):
    pass
#Button 'Button_4' 's Command Event :
def Button_4_onCommand(uiName,widgetName,threadings=0):
    columnTextList = Fun.GetColumnTextList(uiName,'ListView_1',0)
    list_data = []
    for sts in range(0,len(columnTextList)):
        if columnTextList[sts]:
            rowTextList = Fun.GetRowTextList(uiName,'ListView_1',sts)
            list_data.append(rowTextList[1:])
    if len(list_data) == 0:
        Fun.MessageBox("未选择任何数据","提示","info",None) 
        return 
    get_xlsx(list_data)     
#Button 'Button_5' 's Command Event :
def Button_5_onCommand(uiName,widgetName,threadings=0):
    Fun.DeleteAllRows(uiName,'ListView_1')
    pass
#ListView 'ListView_1's HeadingClicked Event :
def ListView_1_onHeadingClicked(uiName,widgetName,columnname,threadings=0):
    if columnname =="☐":
        columnname_state = False
        listviewTextList = Fun.GetAllRowTextList(uiName,'ListView_1')
        for list_data in listviewTextList:
            if list_data[0] == False:
                columnname_state = True
        Fun.SetColumnCheckBox(uiName,'ListView_1',0,-1,0,columnname_state)
